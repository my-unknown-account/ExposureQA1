import os
import re
import json
from collections import Counter
import matplotlib.pyplot as plt
from openpyxl import load_workbook


GROUPS = ["sro", "so", "support"]


def yes_no_to_bool(value):
    if value is None:
        return None

    value = str(value).strip().lower()

    if value == "yes":
        return True
    if value == "no":
        return False

    return None


def parse_sheet_name(sheet_name):
    """
    Example:
    qid_17_support_3 -> group=support, idx=3
    qid_5_sro_1      -> group=sro, idx=1
    """
    m = re.match(r"^(qid_\d+)_(sro|so|support)_(\d+)$", sheet_name)

    if not m:
        return None, None, None

    qid = m.group(1)
    group = m.group(2)
    idx = int(m.group(3))

    return qid, group, idx


def extract_annotations_from_excel_dir(excel_dir):
    annotations = {}

    for file_name in sorted(os.listdir(excel_dir)):
        if not file_name.endswith(".xlsx"):
            continue

        qid = os.path.splitext(file_name)[0]
        file_path = os.path.join(excel_dir, file_name)

        wb = load_workbook(file_path, data_only=True)

        annotations.setdefault(qid, {
            "sro": {},
            "so": {},
            "support": {}
        })

        for sheet_name in wb.sheetnames:
            parsed_qid, group, idx = parse_sheet_name(sheet_name)

            if group is None:
                print(f"Skipping unexpected sheet: {file_name} -> {sheet_name}")
                continue

            ws = wb[sheet_name]
            annotation = yes_no_to_bool(ws["B7"].value)
            relation = ws["B4"].value
            if relation is not None:
                relation = str(relation).strip()
                if not relation:
                    relation = None

            annotations[qid][group][idx] = {
                "annotation": annotation,
                "relation": relation
            }

    return annotations


def majority_vote(values):
    values = [v for v in values if v is not None]

    if not values:
        return None

    counts = Counter(values)

    if counts[True] > counts[False]:
        return True
    if counts[False] > counts[True]:
        return False

    return None


def build_majority_annotations(annotator_dirs):
    all_annotations = [
        extract_annotations_from_excel_dir(d)
        for d in annotator_dirs
    ]

    final_data = {}

    all_qids = set()
    for ann in all_annotations:
        all_qids.update(ann.keys())

    for qid in sorted(all_qids):
        final_data[qid] = {
            "sro": {},
            "so": {},
            "support": {}
        }

        for group in GROUPS:
            all_indices = set()

            for ann in all_annotations:
                if qid in ann:
                    all_indices.update(ann[qid][group].keys())

            for idx in sorted(all_indices):
                votes = []
                relation_votes = []

                for ann in all_annotations:
                    if qid in ann:
                        item = ann[qid][group].get(idx)
                        if item is None:
                            continue
                        votes.append(item.get("annotation"))
                        relation_votes.append(item.get("relation"))

                relation_majority = None
                non_none_relations = [r for r in relation_votes if r is not None]
                if non_none_relations:
                    relation_majority = Counter(non_none_relations).most_common(1)[0][0]

                final_data[qid][group][idx] = {
                    "annotation": majority_vote(votes),
                    "relation": relation_majority
                }

    return final_data


def normalize_samples(samples_file):
    """
    Handles samples.json like:
    {
      "model_a": {"qid_5": ...},
      "model_b": {"qid_5": ...},
      ...
    }

    Keeps the first version of each qid.
    """
    samples = {}

    for model_name, model_data in samples_file.items():
        for qid, item in model_data.items():
            if qid not in samples:
                samples[qid] = item

    return samples


def get_gold_by_index(samples, qid, group, idx):
    items = list(samples[qid][group].items())

    if idx < 1 or idx > len(items):
        return None

    _, value = items[idx - 1]

    if isinstance(value, bool):
        return value

    if isinstance(value, dict):
        return value.get("gold")

    return None


def get_relation_by_index(samples, qid, group, idx):
    items = list(samples[qid][group].items())

    if idx < 1 or idx > len(items):
        return None

    relation, _ = items[idx - 1]
    if relation is None:
        return None

    relation = str(relation).strip()
    return relation if relation else None


def compute_metrics_from_annotations(annotations, samples):
    total_so = 0
    total_sro = 0
    total_support = 0
    true_so = 0
    true_sro = 0
    same_support = 0

    for qid, qid_item in annotations.items():
        if qid not in samples:
            continue
        for group in GROUPS:
            for idx, record in qid_item.get(group, {}).items():
                pred = record.get("annotation")
                if not isinstance(pred, bool):
                    continue

                gold = get_gold_by_index(samples, qid, group, idx)
                if group == "so":
                    total_so += 1
                    if pred:
                        true_so += 1
                if group == "sro":
                    total_sro += 1
                    if pred:
                        true_sro += 1
                if group == "support":
                    total_support += 1
                    if isinstance(gold, bool) and gold == pred:
                        same_support += 1

    return {
        "true_so": true_so,
        "total_so": total_so,
        "true_sro": true_sro,
        "total_sro": total_sro,
        "same_support": same_support,
        "total_support": total_support,
    }


def print_three_metrics(title, metrics):
    so_value = (metrics["true_so"] / metrics["total_so"]) if metrics["total_so"] else 0.0
    sro_value = (metrics["true_sro"] / metrics["total_sro"]) if metrics["total_sro"] else 0.0
    support_value = (metrics["same_support"] / metrics["total_support"]) if metrics["total_support"] else 0.0

    print(title)
    print(f"Lexical SO accuracy: {metrics['true_so']}/{metrics['total_so']} = {so_value:.4f}")
    print(f"Lexical SRO accuracy: {metrics['true_sro']}/{metrics['total_sro']} = {sro_value:.4f}")
    print(f"Relation-Aware Support accuracy: {metrics['same_support']}/{metrics['total_support']} = {support_value:.4f}")


def compute_relation_three_ratios(final_data, samples):
    # For each relation:
    # 1) pred_true / so_all
    # 2) pred_true / sro_all
    # 3) (gold==pred) / support_all
    relation_counts = {}

    for qid, item in final_data.items():
        if qid not in samples:
            continue
        for group in GROUPS:
            for idx, record in item.get(group, {}).items():
                pred = record.get("annotation")
                if not isinstance(pred, bool):
                    continue

                relation = record.get("relation")
                if relation is None:
                    relation = get_relation_by_index(samples, qid, group, idx)
                if relation is None:
                    relation = "Unknown"
                relation = str(relation).strip() or "Unknown"

                if relation not in relation_counts:
                    relation_counts[relation] = {
                        "so_total": 0,
                        "sro_total": 0,
                        "support_total": 0,
                        "true_so": 0,
                        "true_sro": 0,
                        "same_support": 0
                    }

                c = relation_counts[relation]
                if group == "so":
                    c["so_total"] += 1
                    if pred:
                        c["true_so"] += 1
                if group == "sro":
                    c["sro_total"] += 1
                    if pred:
                        c["true_sro"] += 1
                if group == "support":
                    c["support_total"] += 1
                    gold = get_gold_by_index(samples, qid, group, idx)
                    if isinstance(gold, bool) and gold == pred:
                        c["same_support"] += 1

    relation_ratios = {}
    for relation, c in relation_counts.items():
        relation_ratios[relation] = {
            "ratio1_true_over_so": c["true_so"] / c["so_total"] if c["so_total"] else 0.0,
            "ratio2_true_over_sro": c["true_sro"] / c["sro_total"] if c["sro_total"] else 0.0,
            "ratio3_same_over_support": c["same_support"] / c["support_total"] if c["support_total"] else 0.0,
        }
    return relation_ratios


def plot_relation_three_ratios(relation_ratios, output_path):
    if not relation_ratios:
        print("No relation data to plot.")
        return

    relations = sorted(relation_ratios.keys())
    labels = []
    for relation in relations:
        text = str(relation).replace("\n", " ").replace("\t", " ").replace("$", r"\$")
        if len(text) > 80:
            text = text[:77] + "..."
        labels.append(text)

    ratio1 = [relation_ratios[r]["ratio1_true_over_so"] for r in relations]
    ratio2 = [relation_ratios[r]["ratio2_true_over_sro"] for r in relations]
    ratio3 = [relation_ratios[r]["ratio3_same_over_support"] for r in relations]

    x = list(range(len(relations)))
    width = 0.26

    plt.figure(figsize=(max(9, len(relations) * 0.18), 3))
    plt.bar([i - width for i in x], ratio1, width=width, label="Lexical S-O", color="tab:blue")
    plt.bar(x, ratio2, width=width, label="Lexical S-R-O", color="tab:orange")
    plt.bar([i + width for i in x], ratio3, width=width, label="Relation-Aware Support", color="tab:green")

    plt.ylim(0, 1.08)
    plt.ylabel("Accuracy", fontsize=14)
    # plt.xlabel("Relation")
    plt.xticks(x, labels, rotation=45, ha="right", fontsize=14)
    plt.legend(loc="lower center", bbox_to_anchor=(0.5, 1.02), ncol=3, frameon=True, fontsize=14)
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()


def plot_for_person(name, annotations, samples):
    relation_ratios = compute_relation_three_ratios(annotations, samples)
    safe_name = name.lower().replace(" ", "_")
    output_path = f"./three_ratios_by_relation_{safe_name}.pdf"
    plot_relation_three_ratios(relation_ratios, output_path)


if __name__ == "__main__":
    annotator_dirs = [
        "./excel_filled/Jamshid",
        "./excel_filled/Zahra",
        "./excel_filled/Kurosh"
    ]

    with open("./samples.json", "r", encoding="utf-8") as f:
        samples_file = json.load(f)

    samples = normalize_samples(samples_file)

    final_data = build_majority_annotations(annotator_dirs)
    jamshid_annotations = extract_annotations_from_excel_dir("./excel_filled/Jamshid")
    zahra_annotations = extract_annotations_from_excel_dir("./excel_filled/Zahra")
    kurosh_annotations = extract_annotations_from_excel_dir("./excel_filled/Kurosh")

    jamshid_metrics = compute_metrics_from_annotations(jamshid_annotations, samples)
    zahra_metrics = compute_metrics_from_annotations(zahra_annotations, samples)
    kurosh_metrics = compute_metrics_from_annotations(kurosh_annotations, samples)
    final_metrics = compute_metrics_from_annotations(final_data, samples)

    print_three_metrics("Jamshid", jamshid_metrics)
    print()
    print_three_metrics("Zahra", zahra_metrics)
    print()
    print_three_metrics("Kurosh", kurosh_metrics)
    print()
    print_three_metrics("Final", final_metrics)

    plot_for_person("Jamshid", jamshid_annotations, samples)
    plot_for_person("Zahra", zahra_annotations, samples)
    plot_for_person("Kurosh", kurosh_annotations, samples)
    plot_for_person("Final", final_data, samples)
