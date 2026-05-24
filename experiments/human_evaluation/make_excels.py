import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


def make_review_excel(data: dict, output_path="review_items.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    title_fill = PatternFill("solid", fgColor="0F172A")
    label_fill = PatternFill("solid", fgColor="DBEAFE")
    text_fill = PatternFill("solid", fgColor="F8FAFC")
    answer_fill = PatternFill("solid", fgColor="FEF3C7")

    title_font = Font(color="FFFFFF", bold=True, size=14)
    label_font = Font(color="1E3A8A", bold=True)
    text_font = Font(size=11)
    answer_font = Font(color="92400E", bold=True)

    thin_gray = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for qid, item in data.items():
        sub = item.get("sub", "")
        rel = item.get("rel", "")
        obj = item.get("obj", "")

        for group in ["sro", "so", "support"]:
            for idx, (text, _) in enumerate(item.get(group, {}).items(), start=1):
                sheet_name = f"{qid}_{group}_{idx}"[:31]
                ws = wb.create_sheet(title=sheet_name)

                ws.merge_cells("A1:B1")
                ws["A1"] = "Support Annotation"
                ws["A1"].fill = title_fill
                ws["A1"].font = title_font
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

                rows = [
                    ["Subject", sub],
                    ["Relation", rel],
                    ["Object", obj],
                    ["Text", text],
                    ["Is it supported?", ""],
                ]

                start_row = 3

                for r_idx, row in enumerate(rows, start=start_row):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws.cell(r_idx, c_idx, value)
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

                        if c_idx == 1:
                            cell.fill = label_fill
                            cell.font = label_font
                        else:
                            cell.font = text_font
                            cell.fill = text_fill

                ws["A7"].fill = answer_fill
                ws["A7"].font = answer_font
                ws["B7"].fill = answer_fill
                ws["B7"].font = answer_font
                ws["B7"].alignment = Alignment(horizontal="center", vertical="center")

                dv = DataValidation(
                    type="list",
                    formula1='"Yes,No"',
                    allow_blank=False,
                    showErrorMessage=True,
                    errorTitle="Invalid Input",
                    error="Please choose only Yes or No."
                )

                ws.add_data_validation(dv)
                dv.add(ws["B7"])

                ws.column_dimensions["A"].width = 22
                ws.column_dimensions["B"].width = 200

                ws.row_dimensions[1].height = 30
                ws.row_dimensions[3].height = 24
                ws.row_dimensions[4].height = 24
                ws.row_dimensions[5].height = 24
                ws.row_dimensions[6].height = 500
                ws.row_dimensions[7].height = 30

                ws.freeze_panes = "A3"

    wb.save(output_path)


if __name__ == '__main__':
    with open('./samples.json', 'r') as f:
        dataset = json.load(f)
    os.makedirs("excel_files", exist_ok=True)
    for model in dataset:
        for qid, item in dataset[model].items():
            data = {qid: item}
            make_review_excel(data, f"./excel_files/{qid}.xlsx")
    print("Excel file saved as review_items.xlsx")
